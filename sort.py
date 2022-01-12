# --coding:utf-8--
import pandas as pd
import re
from openpyxl import load_workbook, Workbook


class GetDataframe:

    def __init__(self):
        self.all_sheet_data = []

        self.SJ = []  # 砂浆试块
        self.GYC = []  # 钢材
        self.GHJ = []  # 钢焊接
        self.HNT = []  # 混凝土试块

        self.FS = []  # 防水
        self.SN = []  # 水泥
        self.SA = []  # 砂石
        self.SPB = []  # 砼砂配合比
        self.KS = []  # 混凝土抗渗
        self.SZ = []  # 砖试验
        self.TG = []  # 电工套管

    # 获取所有excel为dataframe
    def get_excel(self, file_path):
        df = pd.read_excel(file_path, sheet_name=None)
        for sheet_name in df.keys():
            self.all_sheet_data.append([pd.read_excel(file_path, sheet_name=sheet_name, header=1), sheet_name])
        return self.all_sheet_data

    # 合并基础和主体合并
    def merge_basic_main(self, all_sheet_data):
        # 强制关闭警告
        pd.set_option('mode.chained_assignment', None)
        # 显示所有的列
        pd.set_option('display.max_columns', None)
        for one_sheet_data in all_sheet_data:
            sheet_data = one_sheet_data[0]
            sheet_name = one_sheet_data[1]
            if sheet_name.find('砂浆试块') != -1:
                self.SJ.append(sheet_data)
            elif sheet_name.find('钢材') != -1:
                self.GYC.append(sheet_data)
            elif sheet_name.find('钢筋焊接') != -1:
                self.GHJ.append(sheet_data)
            elif sheet_name.find('混凝土试块') != -1:
                self.HNT.append(sheet_data)

            elif sheet_name.find('防水') != -1:
                self.FS.append(sheet_data)
            elif sheet_name.find('水泥') != -1:
                self.SN.append(sheet_data)
            elif sheet_name.find('砂、石') != -1:
                self.SA.append(sheet_data)
            elif sheet_name.find('砼、砂浆配合比') != -1:
                self.SPB.append(sheet_data)
            elif sheet_name.find('混凝土抗渗') != -1:
                self.KS.append(sheet_data)
            elif sheet_name.find('砖试验') != -1:
                self.SZ.append(sheet_data)
            elif sheet_name.find('电工套管') != -1:
                self.TG.append(sheet_data)
            else:
                print('没找到对应的数据')

        # 砂浆试块,钢筋焊接,混凝土试块需要合并
        for i in range(len(self.SJ) - 1):
            new_SJ = self.SJ[i].append(self.SJ[i + 1])  # append之后都是dataframe类型,相当于两个dataframe合并
        for i in range(len(self.GHJ) - 1):
            new_GHJ = self.GHJ[i].append(self.GHJ[i + 1])
        for i in range(len(self.HNT) - 1):
            new_HNT = self.HNT[i].append(self.HNT[i + 1])

        # dataframe类型
        return {"混凝土试块": new_HNT, "钢筋焊接": new_GHJ, "钢材": self.GYC[0], "砂浆试块": new_SJ, "防水": self.FS[0],
                "水泥": self.SN[0], "砂石": self.SA[0], "砼砂配合比": self.SPB[0], "混凝土抗渗": self.KS[0], "砖试验": self.SZ[0],
                "电工套管": self.TG[0]}


class SortMethod:

    def __init__(self):
        pass

    # 分楼层
    def sort_floor(self, main_body, col_name, sort_date, name):  # 主体数据,楼层所在的列名,排序的日期列,类型例如:混凝土
        temp_data = []  # 每次进行sort_floor,我只需要你返回数据就行了
        # 包含#的主体数据
        body = main_body  # 主体数据
        floor = body[body['{}'.format(col_name)].str.contains('#')]
        # 看共有几个楼层
        all_floor = list(set(floor['{}'.format(col_name)].str[:1].values))
        # 转为int并排序
        all_floor = [int(i) for i in all_floor]
        all_floor = sorted(all_floor)
        # print(all_floor)
        # 循环楼层
        for i in all_floor:
            # print('1sort_floor的i={}---1'.format(i))
            data = floor[floor[col_name].str.startswith('{}#'.format(i))]
            if name == '钢筋焊接':
                result = self.sort_sushe(data, col_name)  # 返回值:已经分号的宿舍和生产车间的数据
                if len(result) > 1:
                    sushe = result[0]
                    sushe = self.sort_zhu_liang(sushe, col_name, sort_date)
                    sushe = self.sort_other(sushe, sort_date)
                    temp_data.append([sushe, '钢筋焊接(主体)宿舍#' + str(i)])

                    chejian = result[1]
                    chejian = self.sort_zhu_liang(chejian, col_name, sort_date)
                    chejian = self.sort_other(chejian, sort_date)
                    temp_data.append([chejian, '钢筋焊接(主体)车间#' + str(i)])
                else:
                    df_sorts = self.sort_zhu_liang(result[0], col_name, sort_date)
                    df_sorts = self.sort_other(df_sorts, sort_date)
                    temp_data.append([df_sorts, '钢筋焊接(主体)#' + str(i)])
            elif name == '砂浆试块':
                df_sorts = self.sort_qiti(data, col_name, sort_date)
                df_sorts = self.sort_other(df_sorts, sort_date)
                temp_data.append([df_sorts, '砂浆试块(主体)#' + str(i)])
            if name == '混凝土试块':
                result = self.sort_sushe(data, col_name)  # 返回值:已经分号的宿舍和生产车间的数据
                if len(result) > 1:
                    sushe = result[0]
                    sushe = self.sort_zhu_liang(sushe, col_name, sort_date)
                    sushe = self.sort_other(sushe, sort_date)
                    temp_data.append([sushe, '混凝土试块(主体)宿舍#' + str(i)])

                    chejian = result[1]
                    chejian = self.sort_zhu_liang(chejian, col_name, sort_date)
                    chejian = self.sort_other(chejian, sort_date)
                    temp_data.append([chejian, '混凝土试块(主体)车间#' + str(i)])
                else:
                    df_sorts = self.sort_zhu_liang(result[0], col_name, sort_date)
                    df_sorts = self.sort_other(df_sorts, sort_date)
                    temp_data.append([df_sorts, '混凝土试块(主体)#' + str(i)])
            else:
                print('floor为None')
        return temp_data

    # 需要分宿舍和生产车间(暂时,有可能有其他,比如综合楼)
    def sort_sushe(self, data, column_name):  # 数据,按照哪一列来分
        mask = data[column_name].str.contains('宿舍', na=True)
        result = len(set(mask.tolist()))  # 判断是否需要分宿舍和生产车间
        if result == 2:  # 需要,因为除了宿舍还有其他的
            sushe = data[data[column_name].str.contains('宿舍')]  # 宿舍
            chejian = data[~data[column_name].str.contains('宿舍')]  # 除了宿舍(暂时只有车间)
            return [sushe, chejian]
        if result == 1:  # 不需要,只有一种(暂时)
            return [data]

    # 报告审编号,序号
    def sort_other(self, data, col_name):  # dataframe,报审表编号按照哪个列排序
        sum = len(data['序号'].values)  # 共有多少条数据,循环的次数
        df = data[col_name].values  # 报审表根据哪个列排序
        # 用来存放最终的报告审编号
        number = []
        temp = 0
        for i in range(sum):
            if i == 0:
                temp = 1
            else:
                if str(df[i]) > str(df[i - 1]):
                    temp += 1
                else:
                    temp = temp
            number.append(temp)
        data['报审表编号'] = number
        # 序号
        ser_number = [{'序号': i + 1} for i in range(len(data.index))]
        number = pd.DataFrame(ser_number)
        # 重置index
        del data['序号']  # 删除原来的序号那一列,用现在排序好的序号列
        data = data.reset_index(drop=True)
        data.insert(0, '序号', number)

        # 如果存在监管登记编号,委托单编号就删除
        if '监管登记编号' and '委托单编号' in data.columns:
            data.drop(['监管登记编号', '委托单编号'], axis=1, inplace=True)
        return data

    # 优先级: zhu_liang,日期,报告编号
    def zhu_liang(self, data):
        if "地坪" in data or "地面板" in data or "地面面层" in data:
            return 1
        elif "一层柱" in data or "一层板" in data or "1层柱" in data:
            return 2
        elif "二层梁" in data:
            return 3
        elif "二层梁板" in data or "二层板" in data:
            return 4
        elif "二层柱" in data or "三层板" in data:
            return 5
        elif "三层梁板" in data:
            return 6
        elif "三层柱" in data:
            return 7
        elif "四层梁板" in data:
            return 8
        elif "四层柱" in data:
            return 9
        elif "五层梁板" in data or "5层梁板" in data:
            return 10
        elif "五层柱" in data:
            return 11
        elif "六层梁板" in data:
            return 12
        elif "六层柱" in data:
            return 13
        elif "屋面" in data or "屋面梁板" in data:
            return 14
        elif "房屋面板" in data:
            return 15
        else:
            return 16

    def sort_zhu_liang(self, data, col_name, col_date):  # 数据,柱和梁所在的列,要排序的日期
        data["临时"] = data[col_name].apply(self.zhu_liang)
        df_sortes = data.sort_values(by=['临时', col_date, '试验报告单编号'], ascending=True)
        df_sortes.drop(['临时'], axis=1, inplace=True)
        return df_sortes

    def qiti(self, data):
        if "一层砌体" in data or "一层" in data:
            return 1
        elif "二层砌体" in data or "二层" in data:
            return 2
        elif "三层砌体" in data or "三层" in data:
            return 3
        elif "四层砌体" in data or "四层" in data:
            return 4
        else:
            return 5

    def sort_qiti(self, data, col_name, col_date):  # 数据,砌体所在的列,要排序的日期
        data["临时"] = data[col_name].apply(self.qiti)
        df_sortes = data.sort_values(by=['临时', col_date, '试验报告单编号'], ascending=True)
        df_sortes.drop(['临时'], axis=1, inplace=True)
        return df_sortes


class MySort:
    def __init__(self):
        self.final_data = []

    def merge(self, data):
        for type in data:
            if type == '混凝土试块':
                self.sort_hnt(data[type])
            elif type == '钢筋焊接':
                self.sort_ghj(data[type])
            elif type == '钢材':
                self.sort_gyc(data[type])
            elif type == '砂浆试块':
                self.sort_sj(data[type])

            elif type == '防水':
                self.sort_fs(data[type])
            elif type == '水泥':
                self.sort_sn(data[type])
            elif type == '砂石':
                self.sort_sa(data[type])
            elif type == '砼砂配合比':
                self.sort_spb(data[type])
            elif type == '混凝土抗渗':
                self.sort_ks(data[type])
            elif type == '砖试验':
                self.sort_sz(data[type])
            elif type == '电工套管':
                self.sort_tg(data[type])
            else:
                pass
        return self.final_data

    def sort_hnt(self, data):  # data为dataframe
        sort_date = '试块成形日期'
        sort_col = '工程部位'
        data['配合比出单日期'] = data[sort_date]
        zhuang = data[data[sort_col].str.contains('桩')]  # 桩数据
        if not zhuang.empty:
            zhuang.sort_values(by=[sort_date, '试验报告单编号'], ascending=True, inplace=True)
            zhuang = SortMethod().sort_other(zhuang, sort_date)
            self.final_data.append([zhuang, '混凝土试块(桩)'])
        # 波浪线代表不包含(不包含桩的数据)
        temp = data[~data[sort_col].str.contains('桩')]
        base = temp[temp[sort_col].str.contains('基础')]  # 基础数据
        if not base.empty:
            base.sort_values(by=[sort_date, '试验报告单编号'], ascending=True, inplace=True)
            base = SortMethod().sort_other(base, sort_date)
            self.final_data.append([base, '混凝土试块(基础)'])
        # temp中不包含基础的就是主体数据
        main_body = temp[~temp[sort_col].str.contains('基础')]  # 主体数据,排序需要传入要根据哪个排序
        if not main_body.empty:
            temp_data = SortMethod().sort_floor(main_body, sort_col, sort_date, '混凝土试块')
            self.final_data.extend(temp_data)
        other = main_body[~main_body[sort_col].str.contains('#')]  # 无楼号的主体数据
        if not other.empty:
            df_sorts = SortMethod().sort_zhu_liang(other, sort_col, sort_date)
            df_sorts = SortMethod().sort_other(df_sorts, sort_date)
            self.final_data.append([df_sorts, '混凝土试块(主体)'])

    def sort_ghj(self, data):  # data为dataframe
        sort_date = '进场日期'
        sort_col = '主要使用部位'
        data.rename(columns={'Unnamed: 2': ''}, inplace=True)
        base = data[data[sort_col].str.contains('基础')]  # 基础数据
        if not base.empty:
            base.sort_values(by=[sort_date, '试验报告单编号'], ascending=True, inplace=True)
            base = SortMethod().sort_other(base, sort_date)
            self.final_data.append([base, '钢筋焊接(基础)'])
        # 不包含基础的就是主体数据
        main_body = data[~data[sort_col].str.contains('基础')]  # 主体数据
        if not main_body.empty:
            temp_data = SortMethod().sort_floor(main_body, sort_col, sort_date, '钢筋焊接')
            self.final_data.extend(temp_data)  # temp_data列表中的数据添加到final_data中
        other = main_body[~main_body[sort_col].str.contains('#')]  # 无楼层的主体数据
        if not other.empty:
            other = SortMethod().sort_zhu_liang(other, sort_col, sort_date)
            self.final_data.append([other, '钢筋焊接(主体)'])

    def sort_gyc(self, data):  # data为dataframe
        sort_date = '进场日期'
        data.rename(columns={'Unnamed: 2': ''}, inplace=True)
        data.iloc[:, 2][data.iloc[:, 2] == 8.0] = '8.0E'
        data.sort_values(by=[sort_date, '试验报告单编号'], ascending=True, inplace=True)
        df = SortMethod().sort_other(data, sort_date)
        self.final_data.append([df, '钢材'])

    def sort_sj(self, data):  # data为dataframe
        sort_date = '试块制作日期'
        sort_col = '工程部位'
        data[sort_date] = data[sort_date].str[:10]
        data['配合比出单日'] = data[sort_date]
        base = data[data[sort_col].str.contains('基础')]  # 基础数据
        if not base.empty:
            base.sort_values(by=[sort_date, '试验报告单编号'], ascending=True, inplace=True)
            base = SortMethod().sort_other(base, sort_date)
            self.final_data.append([base, '砂浆试块(基础)'])
        # temp中不包含基础的就是主体数据
        main_body = data[~data[sort_col].str.contains('基础')]  # 主体数据,排序需要传入要根据哪个排序
        if not main_body.empty:
            temp_data = SortMethod().sort_floor(main_body, sort_col, sort_date, '砂浆试块')
            self.final_data.extend(temp_data)
        temp = main_body[~main_body[sort_col].str.contains('#')]
        two = temp[temp[sort_col].str.contains('±')]  # ±0.0001单独放
        if not two.empty:
            two.sort_values(by=[sort_date, '试验报告单编号'], ascending=True, inplace=True)
            two = SortMethod().sort_other(two, sort_date)
            self.final_data.append([two, '砂浆试块(主体)±0.0001'])
        other = temp[~temp[sort_col].str.contains('±')]  # 无楼层的主体数据
        if not other.empty:
            other = SortMethod().sort_qiti(other, sort_col, sort_date)
            other = SortMethod().sort_other(other, sort_date)
            self.final_data.append([other, '砂浆试块(主体)'])

    def sort_fs(self, data):
        sort_date = '进场日期'
        if not data.empty:
            data.sort_values(by=[sort_date], ascending=True, inplace=True)
            df = SortMethod().sort_other(data, sort_date)
            self.final_data.append([df, '防水'])

    def sort_sn(self, data):
        sort_date = '出厂日期'
        if not data.empty:
            data.sort_values(by=[sort_date], ascending=True, inplace=True)
            df = SortMethod().sort_other(data, sort_date)
            self.final_data.append([df, '水泥'])

    def sort_sa(self, data):
        sort_date = '进场日期'
        if not data.empty:
            data.sort_values(by=[sort_date], ascending=True, inplace=True)
            df = SortMethod().sort_other(data, sort_date)
            self.final_data.append([df, '砂石'])

    def sort_spb(self, data):
        if not data.empty:
            # 如果存在监管登记编号,委托单编号就删除
            if '监管登记编号' and '委托单编号' in data.columns:
                data.drop(['监管登记编号', '委托单编号'], axis=1, inplace=True)
            self.final_data.append([data, '砂配比'])

    def sort_ks(self, data):
        sort_date = '试块成形日期'
        if not data.empty:
            data[sort_date] = data[sort_date].str[:10]
            data.sort_values(by=[sort_date], ascending=True, inplace=True)
            df = SortMethod().sort_other(data, sort_date)
            self.final_data.append([df, '混凝土抗渗'])

    def sort_sz(self, data):
        sort_date = '进场日期'
        if not data.empty:
            data.sort_values(by=[sort_date], ascending=True, inplace=True)
            df = SortMethod().sort_other(data, sort_date)
            self.final_data.append([df, '砖试验'])

    def sort_tg(self, data):
        sort_date = '进场日期'
        if not data.empty:
            data.sort_values(by=[sort_date], ascending=True, inplace=True)
            df = SortMethod().sort_other(data, sort_date)
            self.final_data.append([df, '电工套管'])


class MySave:
    def __init__(self):
        self.pre_path = 'resource/excel/'
        self.pre_file_name = '(分组排序版)'

    # 写入excel之前的排序

    def save_order_sort(self, final_data):
        for i in range(len(final_data)):
            if '钢材' in final_data[i][1]:
                final_data[i].append(1)
            elif '钢筋焊接(基础)' in final_data[i][1]:
                final_data[i].append(2)
            elif '钢筋焊接(主体)' in final_data[i][1]:
                final_data[i].append(3)
            elif '混凝土试块(桩)' in final_data[i][1]:
                final_data[i].append(4)
            elif '混凝土试块(基础)' in final_data[i][1]:
                final_data[i].append(5)
            elif '混凝土试块(主体)' in final_data[i][1]:
                final_data[i].append(6)
            elif '砂浆试块(基础)' in final_data[i][1]:
                final_data[i].append(7)
            elif '砂浆试块(主体)' in final_data[i][1]:
                final_data[i].append(8)
            else:
                final_data[i].append(9)
        final_data = sorted(final_data, key=lambda x: x[2], reverse=False)
        return final_data

    # 创建一个xlsx文件,以工程编号命名
    def create_excel(self, file_path):
        Workbookwb = Workbook()  # 新建一个xlsx文件
        file_name = file_path.split('/')[-1]
        project_number = re.findall('\[(.*?)\]', file_name)[0]
        global save_path
        save_path = self.pre_path + '{0}{1}.xlsx'.format(project_number, self.pre_file_name)
        Workbookwb.save(save_path)  # 文件路经和文件名
        return save_path

    # dataframe写入excel,有已经存在的xlsx文件
    def write_to_excel(self, file_name, data, sheet_name):  # 文件名,每一条dataframe数据,对应的sheet_name
        # file_name = '《乐清经济开发区公交枢纽站建设项目》[G046211]汇总表.xls'
        book = load_workbook(save_path)
        with pd.ExcelWriter(save_path) as E:
            E.book = book
            E.sheets = dict((ws.title, ws) for ws in book.worksheets)  # 获取文件中已存在的表名，这行直接用，不用修改
            data.to_excel(E, sheet_name=sheet_name, startrow=1, index=False)

    # 第一行的工程名称
    def write_first_cow(self, file_name, sheet_name):
        project_name = re.findall('《(.*)》', file_name)[0]
        wb = load_workbook(save_path)
        ws = wb[sheet_name]
        # 这个是从下表1开始,而不是下表0开始
        ws.cell(row=1, column=1).value = project_name  # 第一行的工程名称
        wb.save(save_path)

    # 获取文件名
    def get_fileName(self, file_path):
        file_name = file_path.split('/')[-1]
        # pre_path = file_path.split('/')[:-1]
        # pre_path = "/".join(pre_path) + '/'
        return file_name

    # 删除多余的sheet
    def del_sheet(self):
        wb = load_workbook(save_path)
        ws1 = wb["Sheet"]
        wb.remove(ws1)
        wb.save(save_path)

    # 外部只需调用此方法就可以了
    def my_save(self, file_path, final_data):  # 第三个参数预留save_path
        final_data = self.save_order_sort(final_data)
        self.create_excel(file_path)
        file_name = self.get_fileName(file_path)
        for i in final_data:
            data = i[0]
            sheet_name = i[1]
            self.write_to_excel(file_name, data, sheet_name)
            self.write_first_cow(file_name, sheet_name)
        self.del_sheet()


if __name__ == '__main__':
    file_path = r'resource/excel/《荆山公学教工宿舍扩建工程》[G069338]汇总表.xlsx'

    get = GetDataframe()
    all_sheet_data = get.get_excel(file_path)
    data = get.merge_basic_main(all_sheet_data)

    sort = MySort()
    final_data = sort.merge(data)

    save = MySave()
    save.my_save(file_path, final_data)
